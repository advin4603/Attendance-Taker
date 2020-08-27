from openpyxl import load_workbook, Workbook
from typing import *
import os
import json

# data dictionary containing sheets and their corresponding student lists.
data: Dict[str, Dict[int, str]] = {}

# if no file settings.json found then set default_settings as settings.json
default_settings: Dict[str, Any] = {"dataPath": "",
                                    "downloadPath": os.getcwd()}

# holds settings.
settings: Dict[str, Any] = {}


def load_settings():
    """Reads settings.json and loads settings.json in settings"""
    global settings
    if not os.path.isfile("settings.json"):
        with open("settings.json", "w") as f:
            json.dump(default_settings, f)
        settings = default_settings.copy()
    else:
        with open("settings.json", "r") as f:
            settings = json.load(f)


def change_settings(key: str, value: Any):
    """Changes settings and settings.json"""
    settings[key] = value
    with open("settings.json", "w") as f:
        json.dump(settings, f)


def load_data(excel_file_path: str) -> None:
    """Load the provided excel workbook to data dict"""
    global data
    workbook = load_workbook(excel_file_path)
    data = {sheet: get_data(workbook, sheet) for sheet in workbook.sheetnames}


def get_data(workbook: Workbook, sheet: str = "Main Sheet") -> Dict[int, str]:
    """
    Gets the first column of the spreadsheet in order and returns a dictionary with key as row number and the value
    of the cell.
    """
    return {i: x[0].value for i, x in enumerate(workbook[sheet].rows, start=1) if x[0].value}
