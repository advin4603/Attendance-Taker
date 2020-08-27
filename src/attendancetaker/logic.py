from openpyxl import load_workbook, Workbook
from typing import *
import os
import json

data: Dict[str, Dict[int, str]] = {}
default_settings: Dict[str, Any] = {"dataPath": "",
                                    "downloadPath": os.getcwd()}
settings: Dict[str, Any] = {}


def load_settings():
    global settings
    if not os.path.isfile("settings.json"):
        with open("settings.json", "w") as f:
            json.dump(default_settings, f)
        settings = default_settings.copy()
    else:
        with open("settings.json", "r") as f:
            settings = json.load(f)


def change_settings(key: str, value: Any):
    settings[key] = value
    with open("settings.json", "w") as f:
        json.dump(settings, f)


def load_data(excel_file_path: str) -> None:
    global data
    workbook = load_workbook(excel_file_path)
    data = {sheet: get_data(workbook, sheet) for sheet in workbook.sheetnames}


def get_data(workbook: Workbook, sheet: str = "Main Sheet") -> Dict[int, str]:
    """
        Gets the first column of the spreadsheet in order and returns a dictionary with key as row number and the value of the cell.
    """
    return {i: x[0].value for i, x in enumerate(workbook[sheet].rows, start=1) if x[0].value}
